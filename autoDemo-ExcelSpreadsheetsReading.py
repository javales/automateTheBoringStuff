# /usr/bin/env python3

##############################################################################
# https://www.udemy.com/course/automate/learn/lecture/3465848?start=0#overview
## Section 14: Excel, Word, and PDF Documents (Lesson 42-45) - only Lesson 42 - Reading Excel Spreadsheets

# Install third-party module openpyxl with pip install openpyxl to read and modify Excel files

import openpyxl
import os

os.chdir('/Proj/study/udemy/automateTheBoringStuff')

workbook1 = openpyxl.load_workbook('example.xlsx')

print('Printing the type of the file example.xlsx with type(workbook1):')
print(type(workbook1))
print('')

#sheet1 = workbook1.get_sheet_by_name('Sheet1') #"DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname])"
sheet1a = workbook1['Sheet1']
print('Printing the type of the file\'s Sheet1 with workbrook1[\'Sheet1\']:')
#print(type(sheet1))
print(type(sheet1a))
print('')

#or, if you don't know sheet names, call them with workbook.get_sheet_names()
#actually... DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
print('Printing sheet names with workbook1.sheetnames:')
#print(workbook1.get_sheet_names()) #DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
print(workbook1.sheetnames)
print('')

#getting cells
cellObject1 = sheet1a['A1']
print('Print the cell value of A1:')
print('Raw: (per video, expected this to output as datetime.datetime(2014, 4, 5, 13, 34, 2)')
print(cellObject1.value)
print('String:')
print(str(cellObject1.value))
print('Or use the shortcut of str(sheet1a[\'A1\'].value:')
print(str(sheet1a['A1'].value))
print('')
print('B1:')
print(sheet1a['B1'].value)
print('')
print('C1:')
print(sheet1a['C1'].value)
print('')

#to get cells using only numbers, use the worksheet cell method sheet.cell()
#note that rows and columns begin at 1 NOT 0
print('Print alternative to sheet1a[\'B1\'].value) using numbers instead (as this: "(sheet1a.cell(row=1,column=2)).value")')
print((sheet1a.cell(row=1,column=2)).value)
print('')