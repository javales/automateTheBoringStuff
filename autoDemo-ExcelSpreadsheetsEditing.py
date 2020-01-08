# /usr/bin/env python3

##############################################################################
# https://www.udemy.com/course/automate/learn/lecture/3465848?start=0#overview
## Section 14: Excel, Word, and PDF Documents (Lesson 42-45) - only Lesson 43 - Editing Excel Spreadsheets

# Install third-party module openpyxl with pip install openpyxl to create and modify Excel files

import openpyxl
import os

#workbookObj = openpyxl.load_workbook() #opens an existing Excel file to read

#Create a workbook object
workbookObj = openpyxl.Workbook() #capital Workbook to create a new Excel file

#this starts off with a single sheet as shown by get_sheet_names()
print('Show new Excel workbook\'s sheet names:')
#print(workbookObj.get_sheet_names()) #DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
print(workbookObj.sheetnames)
print('')

#Create a sheet object
print('Creating a sheet object off the sheet "Sheet" shown above:')
#sheetObj = workbookObj.get_sheet_by_name('Sheet') #based on the sheet name shown earlier; DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
sheetObj = workbookObj['Sheet']
print('Sheet object:')
print(sheetObj)
print('')

#assign values to sheet cells
sheetObj['A1'] = 42
sheetObj['A2'] = 'Hello'
print('Printing cells A1 and A2 populated just now with assignment operator:')
print(sheetObj['A1'].value,sheetObj['A2'].value)

#Above assignments only exist in memory and not saved to drive
#to save to drive, use workbookObj.save()
os.chdir('/Proj/study/udemy/automateTheBoringStuff')
workbookObj.save('exampleSave.xlsx')
print('Saved cells to exampleSave.xlsx')
print('')

#Create another sheet and print sheet names
sheetObj2 = workbookObj.create_sheet()
print('Sheet names after creating a second one:')
print(workbookObj.sheetnames)
print('')

#Change name of the new sheet, save to file 2
sheetObj2.title = 'My New Sheet Name'
print('Sheet names after renaming the second one:')
print(workbookObj.sheetnames)
print('')
workbookObj.save('exampleSave2.xlsx')

#Create third sheet at first sheet position, save to file 3
workbookObj.create_sheet(index=0, title='My New First Position Sheet')
print('Sheet names after creating a new one in position 0:')
print(workbookObj.sheetnames)
print('')
workbookObj.save('exampleSave3.xlsx')